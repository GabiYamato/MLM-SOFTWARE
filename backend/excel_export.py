from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .storage import AnalysisResult


def export_results_to_excel(results: Iterable[AnalysisResult]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'MLI Summary'

    header_font = Font(bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    headers = ['Animal', 'Image #', 'Line #', 'Horizontal Intercepts', 'Vertical Intercepts', 'MLI (µm)']
    worksheet.append(headers)
    for column_index, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_index)
        cell.font = header_font
        cell.alignment = center_alignment

    row_index = 2
    for animal_index, result in enumerate(results, start=1):
        animal_label = f'Animal {animal_index}'
        for image in result.images:
            for line in image.lines:
                worksheet.cell(row=row_index, column=1, value=animal_label)
                worksheet.cell(row=row_index, column=2, value=image.image_number)
                worksheet.cell(row=row_index, column=3, value=line.line_number)
                worksheet.cell(row=row_index, column=4, value=line.horizontal_intercepts)
                worksheet.cell(row=row_index, column=5, value=line.vertical_intercepts)
                worksheet.cell(row=row_index, column=6, value=line.mean_linear_intercept_um)
                row_index += 1

            worksheet.cell(row=row_index, column=1, value=f'{animal_label} - Image {image.image_number} average')
            worksheet.merge_cells(start_row=row_index, start_column=1, end_row=row_index, end_column=3)
            worksheet.cell(row=row_index, column=6, value=image.average_mli_um)
            worksheet.cell(row=row_index, column=6).font = header_font
            row_index += 1

    for column in range(1, 7):
        worksheet.column_dimensions[chr(ord('A') + column - 1)].width = 26

    worksheet.cell(row=row_index + 1, column=1, value=f'Generated: {datetime.utcnow().isoformat()}Z')

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
